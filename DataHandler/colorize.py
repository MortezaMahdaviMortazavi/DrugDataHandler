import openpyxl
import re
import string
from openpyxl.styles import PatternFill
import openpyxl


############################################################### Config ##############################################################
file_path = 'yes_no_1.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active
puncs = string.punctuation
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
#####################################################################################################################################

def highlight_cells_boolean_case(col_name, col_data,col_idx,value_to_check, fill_color):
    if not col_name:
        return
    for cell_idx, cell_value in enumerate(col_data, start=1):
        if cell_idx == 2:
            continue
        if cell_value != value_to_check:
            cell = ws.cell(row=cell_idx, column=col_idx)
            cell.fill = fill_color


def extract_equality(s):
    s = s.strip()
    if ' <= ' in s:
        parts = s.split(' <= ')
        operator = '<='
    elif ' >= ' in s:
        parts = s.split(' >= ')
        operator = '>='
    elif ' >=' in s:
        parts = s.split(' >=')
        operator = '>='

    elif ' > ' in s:
        parts = s.split(' > ')
        operator = '>'
    elif ' < ' in s:
        parts = s.split(' < ')
        operator = '<'
    elif '==' in s:
        parts = s.split('==')
        operator = '=='
    else:
        return None

    threshold = eval(str(parts[1]).replace('^', '**')) # Evaluate the expression
    if not isinstance(threshold, (int, float)):
        raise ValueError("Expression result is not a number")

    return operator, threshold


def extract_range(s):
    match = re.search(r'\[(-?\d+(\.\d+)?),(-?\d+(\.\d+)?)\]', s)
    if match:
        return float(match.group(1)), float(match.group(3))
    return None



def processor():
    for col_idx, column in enumerate(ws.iter_cols(), start=1):
        col_data = [cell.value for cell in column]
        col_name = column[0].value
        if col_name is None:
            continue
        # print("col name is",col_name)


        if col_name is None or 'Unnamed' in col_name:
            if col_name is None:
                for cell_idx, cell_value in enumerate(col_data, start=1):
                    if cell_idx == 0:
                        continue

                    if cell_value in ['Yes','Rejected']:
                        cell = ws.cell(row=cell_idx, column=col_idx)
                        cell.fill = red_fill
                continue
            print(f"Skipping column {col_idx}")
            continue
        
        ############################################### Check boolean Cases like Yes or No , Rejected or Accepted #########################
        if col_name in ['Yes', 'No']:
            highlight_cells_boolean_case(
                col_name=col_name,
                col_data=col_data,
                col_idx=col_idx,
                value_to_check=col_name,
                fill_color=red_fill
            )
        ###################################################################################################################################
        
        if 'x' in col_name:
            try:
                operator, threshold = extract_equality(col_name)
            except:
                print("The col that we want extract its x inequallity is",col_name)
                raise ValueError("Something is wrong in extracting the x inequallity Please check")


            if operator is None:
                print(f"Data x is None for column {col_name}")
                continue
            
            for cell_idx, cell_value in enumerate(col_data, start=1):
                if cell_value is not None:
                    try:
                        cell_value = float(str(cell_value).replace("%",""))
                    except (ValueError, TypeError):
                        continue
                
                # print(operator,cell_value,threshold)
                try:
                    if (operator == '>' and cell_value <= threshold) or \
                    (operator == '<' and cell_value >= threshold) or \
                    (operator == '>=' and cell_value < threshold) or \
                    (operator == '<=' and cell_value > threshold) or \
                    (operator == '==' and cell_value != threshold):
                        cell = ws.cell(row=cell_idx, column=col_idx)
                        cell.fill = red_fill

                except:
                    pass


        else:
            
            try:
                col_range = extract_range(col_name)
            except:
                print("Check the column",col_name)
                raise ValueError("Something wrong in extracting the range of coles")
                
            # print(col_range)
            
            for cell_idx, cell_value in enumerate(col_data, start=1):
                if cell_idx == 0:
                    continue
                
                if cell_value in ['Rejected']:
                    cell = ws.cell(row=cell_idx, column=col_idx)
                    cell.fill = red_fill
                

                if cell_value is not None:
                    try:
                        cell_value = float(cell_value)
                    except (ValueError, TypeError):
                        continue
                    
                    # print(col_name,col_range,cell_value,col_range)
                    try:
                        if not (col_range[0] <= cell_value <= col_range[1]):
                            cell = ws.cell(row=cell_idx, column=col_idx)
                            cell.fill = red_fill
                    except:
                        continue

processor()
# # Save the modified workbook
modified_file_path = f'colorize_yes_no2.xlsx'
wb.save(modified_file_path)
wb.close()
print("Cell highlighting completed.")
