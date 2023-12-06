import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook

def greenize_rows_with_red(ws):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Specify the red fill color

    for row in ws.iter_rows(min_row=2):
        has_red = any(cell.fill.start_color.rgb == red_fill.start_color.rgb for cell in row)
        if not has_red:
            row[0].fill = green_fill


def seperate_rows_with_red(ws):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Specify the red fill color

    green_rows = []

    for row in ws.iter_rows(min_row=2):
        has_red = any(cell.fill.start_color.rgb == red_fill.start_color.rgb for cell in row)
        if not has_red:
            row[0].fill = green_fill
            green_rows.append(row)

    return green_rows


file_path = 'test2.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active
greenize_rows_with_red(ws)
wb.save('greenize_red_rows_file.xlsx')

green_rows = seperate_rows_with_red(ws)
new_wb = Workbook()
new_ws = new_wb.active
for row in green_rows:
    new_ws.append([cell.value for cell in row])
new_wb.save('new_file_only_greens.xlsx')