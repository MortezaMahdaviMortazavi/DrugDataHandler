import openpyxl
import re
from openpyxl.styles import PatternFill

# Load the Excel file
file_path = 'dataset/admetlab.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
def extract_equality(s):
    s = s.strip()
    if ' > ' in s:
        parts = s.split(' > ')
        operator = '>'
    elif ' < ' in s:
        parts = s.split(' < ')
        operator = '<'
    elif ' <= ' in s:
        parts = s.split(' <= ')
        operator = '<='
    elif ' >= ' in s:
        parts = s.split(' >= ')
        operator = '>='
    else:
        return None

    threshold = eval(parts[1].replace('^', '**'))  # Evaluate the expression
    if not isinstance(threshold, (int, float)):
        raise ValueError("Expression result is not a number")

    return operator, threshold

def extract_range(s):
    match = re.search(r'\[(-?\d+(\.\d+)?),(-?\d+(\.\d+)?)\]', s)
    if match:
        return float(match.group(1)), float(match.group(3))
    return None

for col_idx, column in enumerate(ws.iter_cols(), start=1):
    col_data = [cell.value for cell in column]
    col_name = column[0].value
    if col_name is None or 'Unnamed' in col_name:
        if col_name is None:
        
            idx = 0
            for cell_idx, cell_value in enumerate(col_data, start=1):
                if idx == 0:
                    idx += 1
                    continue

                if cell_value == 'Rejected':
                    cell = ws.cell(row=cell_idx, column=col_idx)
                    cell.fill = red_fill
            continue
        print(f"Skipping column {col_idx}")
        continue
    
    if 'x' in col_name:
        operator, threshold = extract_equality(col_name)
        if operator is None:
            print(f"Data x is None for column {col_name}")
            continue
        
        for cell_idx, cell_value in enumerate(col_data, start=1):
            if cell_value is not None:
                try:
                    cell_value = float(cell_value)
                except (ValueError, TypeError):
                    continue
                
                if (operator == '>' and cell_value < threshold) or \
                   (operator == '<' and cell_value > threshold) or \
                   (operator == '>=' and cell_value <= threshold) or \
                   (operator == '<=' and cell_value >= threshold):
                    cell = ws.cell(row=cell_idx, column=col_idx)
                    cell.fill = red_fill


    else:
        col_range = extract_range(col_name)
        
        idx = 0
        for cell_idx, cell_value in enumerate(col_data, start=1):
            if idx == 0:
                idx += 1
                continue
            
            print(cell_value)
            if cell_value == 'Rejected':
                cell = ws.cell(row=cell_idx, column=col_idx)
                cell.fill = red_fill

            if cell_value is not None:
                try:
                    cell_value = float(cell_value)
                except (ValueError, TypeError):
                    continue
                
                # print(col_name,col_range,cell_value,col_range)
                try:
                    if not (col_range[0] <= cell_value <= col_range[1]):
                        cell = ws.cell(row=cell_idx, column=col_idx)
                        cell.fill = red_fill
                except:
                    continue

# # Save the modified workbook
modified_file_path = 'results/admetlab_modified.xlsx'
wb.save(modified_file_path)
wb.close()

print("Cell highlighting completed.")